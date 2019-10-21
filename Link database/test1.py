import requests
import json
import xmltodict
#엑셀을 활용하기 위해서 선언
import openpyxl
from collections import OrderedDict

import firebase_admin
from firebase_admin import credentials, firestore

#크롤링 활용
from bs4 import BeautifulSoup
import urllib.request as req
import re

# K-StartUp 공지사항 가져오기

# 공지사항 api url을 커스텀 용이하게 설정
# 현재 필요없는 부분
def setApiUrl():
    # 내 apiKey
    apiKey = "L7N%2BKcwG6vd3iw8aQ7LTzmxErO4N%2FuAqWhHW25Kx5rFa1648arnkkOFObqEbpD4VjelQgqAFgZuh6cemXTwjfQ%3D%3D"

    # 공지사항 파라미터들
    numOfRows = str(10)
    startPage = str(1)
    pageSize = str(10)
    pageNo = str(1)

    url = "http://openapi.kised.or.kr/openapi/service/rest/ContentsService/getNoticeList?serviceKey=" + apiKey + "&numOfRows=" + numOfRows + "&startPage=" + startPage + "&pageSize=" + pageSize + "&pageNo=" + pageNo
    return url


# 파이어베이스 세팅(반환값을 인스턴스로 갖다 쓰면됨.)
def setFirestore():
    # https://firebase.google.com/docs/firestore/quickstart
    cred = credentials.Certificate("service_account_key.json")
    firebase_admin.initialize_app(cred, {
        "projectId": "link-1af3f"
    })
    db = firestore.client()
    return db


# api에서 받았던 xml파일을 작성가능한 json파일로 변환
def xmlToJson():
    xmlString = requests.get(setApiUrl()).text  # api에서 받은 xml 파일 텍스트
    jsonString = json.dumps(xmltodict.parse(xmlString), indent=4)  # 받은 xml을 json 텍스트로 변경(dict로 인식x)
    jsonData = json.loads(jsonString)  # 딕셔너리로 인식할 수 있게 변환
    return jsonData


# 변환한 json파일을 토대로 파이어스토어에 저장
def setJsonToFirestore():
    db = setFirestore()

    for i in xmlToJson()["response"]["body"]["items"]["item"]:
        doc_ref = db.collection("test_공지사항").document()  # 저장하고자 하는 위치
        doc_ref.set({
            "uuid": doc_ref.id,  # 나중에 재접근 용이하게 하기 위해 uuid를 필드에 저장
            "detailurl": i["detailurl"],
            "insertdate": i["insertdate"],
            "postsn": i["postsn"],
            "title": i["title"],
            "viewcount": i["viewcount"]
        })

###### 승균 작성
#uos공지를 크롤링해 Json파일로 저장
def crawlingUOS():
    url = "https://www.uos.ac.kr/korNotice/list.do?list_id=FA1&epTicket=LOG"
    # urlopen()으로 데이터를 가져온다.
    res = req.urlopen(url);
    soup = BeautifulSoup(res, "html.parser");

    # 지속 공지 파트
    normal_list = soup.select("#contents > ul > li > a")

    # 그냥 추가적으로 긁어오기
    link_list = soup.find_all(href=re.compile("^javascript:fnView"))
    link_list2 = soup.find_all(onclikc=re.compile("^fnView"))

    # a_list = soup.find_all("a", {"href": "#"})
    # span_list = soup.find_all("span", {"class": "mhide"})

    for a in link_list:
        print("text: ", a.text)

    for a in link_list2:
        print("text: ", a.text)

    # for a in normal_list:
    #     print("text: ", a.text)
    #
    # for a in a_list:
    #     print("text: ", a.text)
    #
    # for a in span_list:
    #     print("link: ", a.text)

#엑셀 파일을 읽어 Json형태로 변환
def xlsxToJson():
    filename = "공지사항 샘플링(30개).xlsx"
    book = openpyxl.load_workbook(filename)
    sheet = book.worksheets[0]

    #시트의 각 행을 순서대로 추출
    data = []
    for row in range(3, sheet.max_row + 1):
        data_list = OrderedDict()
        data_list['post_id'] = sheet.cell(row, 1).value
        data_list['title'] = sheet.cell(row, 2).value
        data_list['url'] = sheet.cell(row, 7).value
        data.append(data_list)

    jsonString = json.dumps(data, indent=4)  # 받은 xml을 json 텍스트로 변경(dict로 인식x)

    #data.json작성
    with open('data.json', 'w+') as f:
        f.write(jsonString)

    jsonData = json.loads(jsonString)  # 딕셔너리로 인식할 수 있게 변환
    return jsonData




    # 추가적인 활용
    # # 필요없는 줄(헤더, 연도, 계) 제거하기
    # del data[0]
    # #데이터를 정렬하기
    # data = sorted(data, key=lambda x: x[1], reverse=True)
    # #상위 5위를 출력하기
    # for i, a in enumerate(data):
    #     if (i >= 5): break
    #     print(i+1, a[0], int(a[1]))

#엑셀 파일 작성
def jsonToXlsx():
    filename = "test_공지사항.xlsx"
    book = openpyxl.load_workbook(filename)
    # 열려있는 시트에 접근
    # sheet = book.active

    # 맨 앞의 시트 추출
    sheet = book.worksheets[0]

    #시트의 각 행을 순서대로 추출
    data = []
    for row in range(3, sheet.max_row + 1):
        res = 0
        for colume in range(7, 17):
            res += sheet.cell(row, colume).value
        sheet.cell(row, 18).value = res/10

    book.save('./new_test.xlsx')

# 변환한 json파일을 토대로 파이어스토어에 저장
def setXlsxToFirestore():
    db = setFirestore()
    # json파일만 활용하기 위해서
    # with open('data.json') as Json_file:

    for i in xlsxToJson():
        doc_ref = db.collection("test_공지사항").document()  # 저장하고자 하는 위치
        doc_ref.set({
            "uuid": doc_ref.id,  # 나중에 재접근 용이하게 하기 위해 uuid를 필드에 저장
            "detailurl": i["url"],
            "title": i["title"],
            "post_id": i["post_id"]
        })


if __name__ == '__main__':
    # setJsonToFirestore()
    # xlsxToJson()
    # jsonToXlsx()
    #crawlingUOS()
    setXlsxToFirestore()